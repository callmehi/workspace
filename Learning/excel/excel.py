import openpyxl

# wb=openpyxl.load_workbook('excel.xlsx')

# wb.create_sheet("Report", 1) # 新增工作表並指定放置位置
# wb.create_sheet("Data", 0) # 新增工作表並指定放置位置

# ws=wb.actives

# ws['A2'].value = 'emily' 
# ws['A3'].value = 'emily'
# ws['A4'].value = 'emily'
# ws['A5'].value = 'emily'
# ws['A6'].value = 'emily'

fn = 'excel.xlsx'
wb = openpyxl.load_workbook(fn)

print(wb.sheetnames)    # 讀取excel檔案每個工作表的名稱
print(wb.active)        # 讀取excel檔案每個工作表的名稱
print(wb.active.title)  # 讀取excel檔案每個工作表的名稱

sheet = wb['Data1']
sheet.title = 'Pear'
wb.save("excel.xlsx")


# wb.save('excel.xlsx')