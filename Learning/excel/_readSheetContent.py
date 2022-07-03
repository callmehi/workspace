import openpyxl

# fn = 'new_excel.xlsx'
# wb = openpyxl.load_workbook(fn)

# wb.active = 0
# ws = wb.active
# print('excel活動工作表： ', ws)


# for row in ws:
#     for cell in row:
#         print(cell.value)

# print()
# print('D1內容： ', ws['D1'].value)


fn = 'new_excel.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active

range = ws['A1': 'C10']

for a, b, c  in range:
    print("{0} {1} {2}".format(a.value, b.value, c.value))

print()
# for a, b, c in ws[ws.dimensions]:
#     print(a.value, b.value, c.value)