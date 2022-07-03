import openpyxl

fn = 'new_excel.xlsx'
wb = openpyxl.load_workbook(fn)

print('excel活動工作表： ', wb.active)

wb.active = 2
print('excel活動工作表： ', wb.active)
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

fn = 'new_excel.xlsx'
wb = openpyxl.load_workbook(fn)

print('目前工作表： ', wb.active.title)

wb.active = wb['Sheet']
print('目前工作表： ', wb.active.title)